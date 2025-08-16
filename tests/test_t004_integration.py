"""
Integration tests for T-004: Provenance JSON + embed meta in Excel
"""
import pytest
import tempfile
import json
from pathlib import Path
from datetime import datetime

from src.core.provenance import build_run_meta, write_run_json
from src.core.provenance_integration import (
    load_provenance_from_run_dir,
    create_excel_metadata,
    get_provenance_summary,
    create_provenance_for_sync_run
)
from src.core.gdx_io_merg import export_excel
import pandas as pd


class TestProvenanceGeneration:
    """Test provenance JSON generation"""
    
    def test_build_run_meta(self):
        """Test build_run_meta creates proper metadata"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a minimal model directory
            model_dir = Path(tmpdir) / "model"
            model_dir.mkdir()
            (model_dir / "main.gms").write_text("$ontext test model $offtext")
            
            meta = build_run_meta(
                work_dir=str(model_dir),
                main_file="main.gms",
                options={"Lo": "2"},
                scenario_id="test_scenario"
            )
            
            # Check required fields
            assert "run_id" in meta
            assert "timestamp" in meta
            assert "model_hash" in meta
            assert meta["main_file"] == "main.gms"
            assert meta["options"] == {"Lo": "2"}
            assert meta["scenario_id"] == "test_scenario"
            
            # Check timestamp format
            assert "T" in meta["timestamp"]
            assert meta["timestamp"].endswith("Z")
    
    def test_write_run_json(self):
        """Test write_run_json creates proper JSON file"""
        with tempfile.TemporaryDirectory() as tmpdir:
            run_dir = Path(tmpdir)
            
            meta = {
                "run_id": "test-123",
                "timestamp": "2025-08-16T20:00:00Z",
                "model_hash": "abcdef123456"
            }
            
            json_path = write_run_json(run_dir, meta)
            
            assert json_path.exists()
            assert json_path.name == "run.json"
            
            # Verify content
            loaded_data = json.loads(json_path.read_text())
            assert loaded_data["run_id"] == "test-123"
            assert loaded_data["timestamp"] == "2025-08-16T20:00:00Z"
            assert loaded_data["model_hash"] == "abcdef123456"


class TestProvenanceIntegration:
    """Test provenance integration utilities"""
    
    def test_load_provenance_from_run_dir(self):
        """Test loading provenance data from run directory"""
        with tempfile.TemporaryDirectory() as tmpdir:
            run_dir = Path(tmpdir)
            
            # No run.json - should return None
            result = load_provenance_from_run_dir(run_dir)
            assert result is None
            
            # Create run.json
            run_json = run_dir / "run.json"
            test_data = {"run_id": "test-456", "timestamp": "2025-08-16T20:00:00Z"}
            run_json.write_text(json.dumps(test_data))
            
            # Should load successfully
            result = load_provenance_from_run_dir(run_dir)
            assert result is not None
            assert result["run_id"] == "test-456"
            assert result["timestamp"] == "2025-08-16T20:00:00Z"
    
    def test_create_excel_metadata(self):
        """Test Excel metadata creation"""
        # Test without provenance data
        meta = create_excel_metadata(
            run_id="ui-123",
            gdx_file="test.gdx",
            units={"power": "MW"},
            symbols_count=5,
            symbols_with_data=3,
            duration_seconds=15.5
        )
        
        assert meta["ui_run_id"] == "ui-123"
        assert meta["gdx_file"] == "test.gdx"
        assert meta["units_applied"] == "power=MW"
        assert meta["symbols_count"] == "5"
        assert meta["symbols_with_data"] == "3"
        assert meta["duration_seconds"] == "15.5"
        assert "export_timestamp" in meta
        
        # Test with provenance data
        provenance_data = {
            "run_id": "prov-789",
            "timestamp": "2025-08-16T20:00:00Z",
            "model_hash": "abc123",
            "main_file": "model.gms"
        }
        
        meta_with_prov = create_excel_metadata(
            provenance_data=provenance_data,
            run_id="ui-123"
        )
        
        assert meta_with_prov["provenance_run_id"] == "prov-789"
        assert meta_with_prov["provenance_timestamp"] == "2025-08-16T20:00:00Z"
        assert meta_with_prov["model_hash"] == "abc123"
        assert meta_with_prov["main_file"] == "model.gms"
        assert meta_with_prov["ui_run_id"] == "ui-123"
    
    def test_get_provenance_summary(self):
        """Test provenance summary generation"""
        with tempfile.TemporaryDirectory() as tmpdir:
            run_dir = Path(tmpdir)
            
            # No run.json
            summary = get_provenance_summary(run_dir)
            assert "status" in summary
            assert summary["status"] == "No provenance data available"
            
            # With run.json
            run_json = run_dir / "run.json"
            test_data = {
                "run_id": "test-summary",
                "timestamp": "2025-08-16T20:00:00Z",
                "model_hash": "abcdef123456789012345678901234567890abcdef123456",
                "main_file": "test.gms",
                "options": {"Lo": "2"},
                "scenario_id": "test_scenario"
            }
            run_json.write_text(json.dumps(test_data))
            
            summary = get_provenance_summary(run_dir)
            assert summary["Run ID"] == "test-summary"
            assert summary["Timestamp"] == "2025-08-16T20:00:00Z"
            assert summary["Model Hash"].endswith("...")  # Should be truncated
            assert summary["Main File"] == "test.gms"
            assert summary["Options"] == "{'Lo': '2'}"
            assert summary["Scenario"] == "test_scenario"
    
    def test_create_provenance_for_sync_run(self):
        """Test provenance creation for sync runs"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a minimal model directory
            model_dir = Path(tmpdir) / "model"
            model_dir.mkdir()
            (model_dir / "sync.gms").write_text("$ontext sync test $offtext")
            
            output_dir = Path(tmpdir) / "output"
            output_dir.mkdir()
            
            meta = create_provenance_for_sync_run(
                work_dir=str(model_dir),
                main_file="sync.gms",
                options={"action": "c"},
                output_dir=output_dir
            )
            
            assert meta is not None
            assert meta["main_file"] == "sync.gms"
            assert meta["scenario_id"] == "sync_run"
            assert "run_id" in meta
            
            # Check that run.json was created
            run_json = output_dir / "run.json"
            assert run_json.exists()
            
            loaded_meta = json.loads(run_json.read_text())
            assert loaded_meta["main_file"] == "sync.gms"


class TestExcelIntegration:
    """Test Excel integration with provenance data"""
    
    def test_excel_export_with_provenance_metadata(self):
        """Test Excel export includes provenance metadata"""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create test data
            symbol_data = {
                "test_param": pd.DataFrame([
                    {"key1": "A", "value": 100.0},
                    {"key1": "B", "value": 200.0}
                ])
            }
            
            # Create provenance metadata
            provenance_meta = create_excel_metadata(
                provenance_data={
                    "run_id": "prov-excel-test",
                    "timestamp": "2025-08-16T20:00:00Z",
                    "model_hash": "excel123",
                    "main_file": "excel_test.gms",
                    "options": {"Lo": "2"}
                },
                run_id="ui-excel-test",
                gdx_file="test.gdx",
                units={"test_param": "MW"},
                symbols_count=1,
                symbols_with_data=1
            )
            
            # Export to Excel
            xlsx_path = Path(tmpdir) / "test_provenance.xlsx"
            result = export_excel(
                symbol_data=symbol_data,
                xlsx_out=xlsx_path,
                units={"test_param": "MW"},
                meta=provenance_meta
            )
            
            assert result.exists()
            assert result.suffix == ".xlsx"
            
            # Verify Excel content (basic check)
            import openpyxl
            wb = openpyxl.load_workbook(result)
            
            # Should have meta sheet
            assert "meta" in wb.sheetnames
            
            # Should have data sheet
            assert "test_param" in wb.sheetnames
            
            # Check meta sheet has provenance data
            meta_sheet = wb["meta"]
            meta_values = {}
            for row in meta_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    meta_values[row[0]] = row[1]
            
            assert "provenance_run_id" in meta_values
            assert meta_values["provenance_run_id"] == "prov-excel-test"
            assert "model_hash" in meta_values
            assert meta_values["model_hash"] == "excel123"
            assert "ui_run_id" in meta_values
            assert meta_values["ui_run_id"] == "ui-excel-test"


class TestEndToEndWorkflow:
    """Test complete T-004 workflow"""
    
    @pytest.mark.skipif(
        not Path("toy_model").exists(),
        reason="toy_model directory not found - skipping integration test"
    )
    def test_complete_t004_workflow(self):
        """Test complete T-004 workflow with real model"""
        # This test would require running an actual GAMS model
        # For now, we'll test the components separately
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # Simulate run directory with provenance
            run_dir = Path(tmpdir) / "run_test"
            run_dir.mkdir()
            
            # Create mock provenance data
            provenance_data = {
                "run_id": "workflow-test-123",
                "timestamp": "2025-08-16T20:00:00Z",
                "model_hash": "workflow123abc",
                "main_file": "main.gms",
                "options": {"Lo": "2"},
                "scenario_id": "integration_test"
            }
            
            write_run_json(run_dir, provenance_data)
            
            # Test loading and summarizing
            loaded = load_provenance_from_run_dir(run_dir)
            assert loaded["run_id"] == "workflow-test-123"
            
            summary = get_provenance_summary(run_dir)
            assert summary["Run ID"] == "workflow-test-123"
            
            # Test Excel metadata creation
            excel_meta = create_excel_metadata(
                provenance_data=loaded,
                run_id="ui-workflow-test",
                gdx_file="results.gdx"
            )
            
            assert excel_meta["provenance_run_id"] == "workflow-test-123"
            assert excel_meta["ui_run_id"] == "ui-workflow-test"
            assert excel_meta["scenario_id"] == "integration_test"
            
            # Test Excel export
            test_data = {
                "workflow_param": pd.DataFrame([{"key1": "test", "value": 42.0}])
            }
            
            xlsx_path = Path(tmpdir) / "workflow_test.xlsx"
            export_excel(
                symbol_data=test_data,
                xlsx_out=xlsx_path,
                meta=excel_meta
            )
            
            assert xlsx_path.exists()
            
            # Verify Excel contains both UI and provenance metadata
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path)
            meta_sheet = wb["meta"]
            
            meta_keys = [cell.value for cell in meta_sheet["A"]]
            assert "provenance_run_id" in meta_keys
            assert "ui_run_id" in meta_keys
            assert "model_hash" in meta_keys